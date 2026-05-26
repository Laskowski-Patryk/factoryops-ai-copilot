from __future__ import annotations

import csv
import re
import shutil
import sqlite3
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path
from uuid import uuid4

from fastapi import UploadFile
from openpyxl import load_workbook

from app.models.schemas import DatasetSummary
from app.tools.factory_tools import assert_select_only

MAX_ROWS_PER_SHEET = 10000
BLOCKED_PYTHON = re.compile(
    r"\b(import|open|exec|eval|compile|input|globals|locals|__import__)\b|__|"
    r"\b(os|sys|subprocess|socket|requests|pathlib|shutil)\b",
    re.I,
)


class DatasetService:
    def __init__(self, uploads_dir: str) -> None:
        self.uploads_dir = Path(uploads_dir).resolve()
        self.uploads_dir.mkdir(parents=True, exist_ok=True)

    def upload(self, file: UploadFile) -> DatasetSummary:
        filename = Path(file.filename or "dataset").name
        suffix = Path(filename).suffix.lower()
        dataset_id = str(uuid4())
        folder = self.uploads_dir / dataset_id
        folder.mkdir(parents=True, exist_ok=True)
        source_path = folder / filename
        with source_path.open("wb") as handle:
            shutil.copyfileobj(file.file, handle)

        db_path = folder / "dataset.sqlite3"
        if suffix == ".csv":
            kind = "csv"
            self._csv_to_sqlite(source_path, db_path)
        elif suffix in {".xlsx", ".xlsm"}:
            kind = "excel"
            self._excel_to_sqlite(source_path, db_path)
        elif suffix in {".sqlite", ".sqlite3", ".db"}:
            kind = "sqlite"
            shutil.copy2(source_path, db_path)
        else:
            raise ValueError("Supported uploads: CSV, XLSX/XLSM, SQLite DB.")

        summary = self._summarize(dataset_id, filename, kind, db_path)
        (folder / "metadata.json").write_text(summary.model_dump_json(indent=2), encoding="utf-8")
        return summary

    def list(self) -> list[DatasetSummary]:
        summaries = []
        for metadata in self.uploads_dir.glob("*/metadata.json"):
            summaries.append(
                DatasetSummary.model_validate_json(metadata.read_text(encoding="utf-8"))
            )
        return sorted(summaries, key=lambda item: item.created_at, reverse=True)

    def get(self, dataset_id: str) -> DatasetSummary | None:
        try:
            metadata = self._dataset_dir(dataset_id) / "metadata.json"
        except ValueError:
            return None
        if not metadata.exists():
            return None
        return DatasetSummary.model_validate_json(metadata.read_text(encoding="utf-8"))

    def delete(self, dataset_id: str) -> bool:
        folder = self._dataset_dir(dataset_id)
        if not folder.exists():
            return False
        shutil.rmtree(folder)
        return True

    def profile(self, dataset_id: str) -> dict:
        db_path = self._db_path(dataset_id)
        with sqlite3.connect(db_path) as conn:
            conn.row_factory = sqlite3.Row
            tables = self._tables(conn)
            table_profiles = []
            for table in tables:
                columns = self._columns(conn, table)
                rows = conn.execute(f'SELECT * FROM "{table}" LIMIT 1000').fetchall()
                row_count = conn.execute(f'SELECT COUNT(*) AS count FROM "{table}"').fetchone()[
                    "count"
                ]
                samples = [dict(row) for row in rows[:5]]
                table_profiles.append(
                    {
                        "table": table,
                        "row_count": row_count,
                        "columns": columns,
                        "sample_rows": samples,
                        "numeric_stats": self._numeric_stats(rows, columns),
                        "top_values": self._top_values(rows, columns),
                        "detected_roles": self._detect_roles(columns),
                    }
                )
        return {"dataset_id": dataset_id, "tables": table_profiles}

    def query(self, dataset_id: str, sql: str) -> list[dict]:
        assert_select_only(sql)
        with sqlite3.connect(self._db_path(dataset_id)) as conn:
            conn.row_factory = sqlite3.Row
            return [dict(row) for row in conn.execute(sql).fetchmany(100)]

    def run_python_analysis(self, dataset_id: str, code: str) -> dict:
        if BLOCKED_PYTHON.search(code):
            raise ValueError("Python analysis code contains blocked operations.")
        profile = self.profile(dataset_id)
        table = profile["tables"][0]["table"]
        rows = self.query(dataset_id, f'SELECT * FROM "{table}" LIMIT 500')
        namespace = {
            "rows": rows,
            "result": None,
            "len": len,
            "sum": sum,
            "min": min,
            "max": max,
            "round": round,
            "sorted": sorted,
        }
        exec(code, {"__builtins__": {}}, namespace)
        return {
            "table": table,
            "rows_available": len(rows),
            "result": namespace.get("result"),
        }

    def infer_factory_metrics(self, dataset_id: str) -> dict:
        profile = self.profile(dataset_id)
        best = profile["tables"][0] if profile["tables"] else None
        if not best:
            return {"mode": "empty", "cards": [], "insights": ["Uploaded database has no tables."]}

        roles = best["detected_roles"]
        cards = [{"label": "Rows", "value": str(best["row_count"]), "detail": best["table"]}]
        insights = [
            f"Detected table `{best['table']}` with {best['row_count']} rows and "
            f"{len(best['columns'])} columns."
        ]
        if roles.get("line"):
            insights.append(f"Line/group column candidate: `{roles['line']}`.")
        if roles.get("date"):
            insights.append(f"Date/time column candidate: `{roles['date']}`.")

        stats = best["numeric_stats"]
        for role_name, label in [
            ("target", "Target"),
            ("actual", "Actual"),
            ("oee", "OEE"),
            ("downtime", "Downtime"),
            ("scrap", "Scrap"),
        ]:
            column = roles.get(role_name)
            if column and column in stats:
                stat = stats[column]
                value = stat["latest"] or stat["avg"]
                cards.append(
                    {
                        "label": label,
                        "value": _format_number(value),
                        "detail": f"column `{column}` avg {_format_number(stat['avg'])}",
                    }
                )

        return {
            "mode": "uploaded_dataset",
            "table": best["table"],
            "roles": roles,
            "cards": cards[:8],
            "insights": insights,
            "profile": profile,
        }

    def _csv_to_sqlite(self, source_path: Path, db_path: Path) -> None:
        with source_path.open("r", encoding="utf-8-sig", newline="") as handle:
            sample = handle.read(4096)
            handle.seek(0)
            dialect = csv.Sniffer().sniff(sample) if sample else csv.excel
            reader = csv.reader(handle, dialect)
            headers = next(reader, [])
            table = "uploaded_data"
            self._rows_to_sqlite(db_path, table, headers, reader)

    def _excel_to_sqlite(self, source_path: Path, db_path: Path) -> None:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            headers = ["" if value is None else str(value) for value in next(rows, [])]
            self._rows_to_sqlite(db_path, _safe_name(sheet.title), headers, rows)

    def _rows_to_sqlite(self, db_path: Path, table: str, headers: list, rows) -> None:
        columns = _dedupe(
            [_safe_name(str(header or f"column_{idx + 1}")) for idx, header in enumerate(headers)]
        )
        if not columns:
            columns = ["value"]
        with sqlite3.connect(db_path) as conn:
            conn.execute(f'DROP TABLE IF EXISTS "{table}"')
            conn.execute(
                f'CREATE TABLE "{table}" ({", ".join(f"""\"{col}\" TEXT""" for col in columns)})'
            )
            placeholders = ", ".join("?" for _ in columns)
            for idx, row in enumerate(rows):
                if idx >= MAX_ROWS_PER_SHEET:
                    break
                values = list(row)
                values = values[: len(columns)] + [""] * max(0, len(columns) - len(values))
                conn.execute(
                    f'INSERT INTO "{table}" VALUES ({placeholders})',
                    ["" if value is None else str(value) for value in values],
                )

    def _summarize(
        self, dataset_id: str, filename: str, kind: str, db_path: Path
    ) -> DatasetSummary:
        with sqlite3.connect(db_path) as conn:
            tables = self._tables(conn)
            row_count = 0
            columns = []
            for table in tables:
                row_count += conn.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]
                columns.extend(self._columns(conn, table))
        return DatasetSummary(
            id=dataset_id,
            name=filename,
            kind=kind,
            created_at=datetime.now(UTC),
            tables=tables,
            row_count=row_count,
            columns=sorted(set(columns)),
        )

    def _db_path(self, dataset_id: str) -> Path:
        path = self._dataset_dir(dataset_id) / "dataset.sqlite3"
        if not path.exists():
            raise ValueError(f"Dataset not found: {dataset_id}")
        return path

    def _dataset_dir(self, dataset_id: str) -> Path:
        if not re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9_.-]{0,100}", dataset_id):
            raise ValueError("Invalid dataset id")
        path = (self.uploads_dir / dataset_id).resolve()
        if self.uploads_dir != path and self.uploads_dir not in path.parents:
            raise ValueError("Invalid dataset path")
        return path

    def _tables(self, conn: sqlite3.Connection) -> list[str]:
        rows = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
        ).fetchall()
        return [row[0] for row in rows]

    def _columns(self, conn: sqlite3.Connection, table: str) -> list[str]:
        return [row[1] for row in conn.execute(f'PRAGMA table_info("{table}")').fetchall()]

    def _numeric_stats(self, rows: list[sqlite3.Row], columns: list[str]) -> dict:
        stats = {}
        for column in columns:
            values = [_to_float(row[column]) for row in rows]
            numeric = [value for value in values if value is not None]
            if len(numeric) >= max(2, len(rows) * 0.4):
                stats[column] = {
                    "min": min(numeric),
                    "max": max(numeric),
                    "avg": round(sum(numeric) / len(numeric), 3),
                    "latest": numeric[-1],
                }
        return stats

    def _top_values(self, rows: list[sqlite3.Row], columns: list[str]) -> dict:
        top_values = {}
        for column in columns:
            values = [str(row[column]) for row in rows if row[column] not in {None, ""}]
            if values:
                top_values[column] = Counter(values).most_common(5)
        return top_values

    def _detect_roles(self, columns: list[str]) -> dict[str, str]:
        role_patterns = {
            "line": ["line", "cell", "machine", "asset", "workcenter", "work_center"],
            "date": ["date", "day", "timestamp", "time", "shift"],
            "target": ["target", "plan", "planned", "quota"],
            "actual": ["actual", "output", "produced", "units", "quantity", "qty"],
            "oee": ["oee", "efficiency"],
            "scrap": ["scrap", "reject", "defect"],
            "downtime": ["downtime", "stop", "loss_minutes", "minutes_down"],
            "reason": ["reason", "fault", "cause", "category"],
        }
        roles = {}
        normalized = {column: column.lower().replace(" ", "_") for column in columns}
        for role, patterns in role_patterns.items():
            for column, value in normalized.items():
                if any(pattern in value for pattern in patterns):
                    roles[role] = column
                    break
        return roles


def _safe_name(value: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9_]+", "_", value.strip().lower()).strip("_")
    return cleaned or "column"


def _dedupe(values: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    output = []
    for value in values:
        count = seen.get(value, 0)
        seen[value] = count + 1
        output.append(value if count == 0 else f"{value}_{count + 1}")
    return output


def _to_float(value) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", ".").replace("%", "")
    if not text:
        return None
    try:
        number = float(text)
        return number / 100 if "%" in str(value) and number > 1 else number
    except ValueError:
        return None


def _format_number(value: float | int | None) -> str:
    if value is None:
        return "n/a"
    if abs(float(value)) >= 100:
        return f"{float(value):,.0f}"
    return f"{float(value):.2f}".rstrip("0").rstrip(".")
